"""Stats calculator ve Excel writer testleri."""

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook
import io

from models.schemas import (
    ExamStats,
    GradeResult,
    QuestionDetail,
    ScanResult,
    StudentAnswer,
    AnswerKey,
)
from services.stats_calculator import calculate_stats
from services.excel_writer import create_excel
from services.grader import grade_exam


# ---------------------------------------------------------------------------
# Yardimci: ornek veri uret
# ---------------------------------------------------------------------------


def _make_grade_result(
    sinav_id: str,
    grup: str,
    ogrenci_ad: str,
    cevaplar: list[str | None],
    dogru_cevaplar: list[str],
    puanlar: list[float],
) -> GradeResult:
    """Scan + answer key'den GradeResult olustur."""
    scan = ScanResult(
        sinav_id=sinav_id,
        grup=grup,
        ogrenci_ad=ogrenci_ad,
        ogrenci_no=None,
        soru_sayisi=len(cevaplar),
        sik_sayisi=4,
        cevaplar=[
            StudentAnswer(soru_no=i + 1, cevap=c, gecersiz=False)
            for i, c in enumerate(cevaplar)
        ],
    )
    key = AnswerKey(
        sinav_id=sinav_id,
        grup=grup,
        cevaplar=dogru_cevaplar,
        puanlar=puanlar,
    )
    return grade_exam(scan, key)


# ---------------------------------------------------------------------------
# Stats calculator testleri
# ---------------------------------------------------------------------------


class TestStatsCalculator:
    def test_empty_results(self):
        stats = calculate_stats("test1", [])
        assert stats.toplam_ogrenci == 0
        assert stats.sinif_ortalamasi == 0.0
        assert stats.soru_dogru_oranlari == []

    def test_single_student(self):
        r = _make_grade_result(
            "test1",
            "A",
            "Ali",
            ["A", "B", "C", "D", None],
            ["A", "B", "A", "D", "C"],
            [20, 20, 20, 20, 20],
        )
        stats = calculate_stats("test1", [r])

        assert stats.toplam_ogrenci == 1
        assert stats.sinif_ortalamasi == 60.0
        assert stats.medyan == 60.0
        assert stats.en_yuksek == 60.0
        assert stats.en_dusuk == 60.0
        assert stats.standart_sapma == 0.0
        assert len(stats.soru_dogru_oranlari) == 5
        assert stats.soru_dogru_oranlari[0] == 1.0  # S1 dogru
        assert stats.soru_dogru_oranlari[2] == 0.0  # S3 yanlis
        assert stats.soru_dogru_oranlari[4] == 0.0  # S5 bos

    def test_multiple_students_multiple_groups(self):
        dogru = ["A", "B", "C", "D"]
        puanlar = [25, 25, 25, 25]

        r1 = _make_grade_result("exam1", "A", "Ali", ["A", "B", "C", "D"], dogru, puanlar)
        r2 = _make_grade_result("exam1", "A", "Veli", ["A", "B", "A", "D"], dogru, puanlar)
        r3 = _make_grade_result("exam1", "B", "Ayse", ["A", "A", "C", "A"], dogru, puanlar)

        stats = calculate_stats("exam1", [r1, r2, r3])

        assert stats.toplam_ogrenci == 3
        assert stats.en_yuksek == 100.0
        assert stats.en_dusuk == 50.0
        assert len(stats.grup_ortalamalari) == 2
        assert "A" in stats.grup_ortalamalari
        assert "B" in stats.grup_ortalamalari
        # Grup A: (100 + 75) / 2 = 87.5
        assert stats.grup_ortalamalari["A"] == 87.5
        # Grup B: 50 / 1 = 50.0
        assert stats.grup_ortalamalari["B"] == 50.0
        # En zor sorular (en az dogru orani olan)
        assert len(stats.en_zor_sorular) == 4

    def test_hardest_questions_sorted(self):
        dogru = ["A", "B", "C", "D", "A"]
        puanlar = [20, 20, 20, 20, 20]

        # 3 ogrenci, S3 hicbir ogrenci dogru yapmaz
        r1 = _make_grade_result("e1", "A", "O1", ["A", "B", "D", "D", "A"], dogru, puanlar)
        r2 = _make_grade_result("e1", "A", "O2", ["A", "B", "A", "D", "B"], dogru, puanlar)
        r3 = _make_grade_result("e1", "A", "O3", ["A", "A", "B", "A", "A"], dogru, puanlar)

        stats = calculate_stats("e1", [r1, r2, r3])
        # S3 (idx 2) = 0/3 = 0.0, en zor
        assert stats.en_zor_sorular[0] == 3


# ---------------------------------------------------------------------------
# Excel writer testleri
# ---------------------------------------------------------------------------


class TestExcelWriter:
    def test_creates_two_sheets(self):
        dogru = ["A", "B", "C"]
        puanlar = [10, 10, 10]

        r1 = _make_grade_result("ex1", "A", "Ali", ["A", "B", "C"], dogru, puanlar)
        r2 = _make_grade_result("ex1", "A", "Veli", ["A", "A", "C"], dogru, puanlar)

        stats = calculate_stats("ex1", [r1, r2])
        xlsx_bytes = create_excel("ex1", "Matematik Vize", [r1, r2], stats)

        wb = load_workbook(io.BytesIO(xlsx_bytes))
        assert len(wb.sheetnames) == 2
        assert wb.sheetnames[0] == "Sonuclar"
        assert wb.sheetnames[1] == "Istatistikler"

    def test_results_sheet_content(self):
        dogru = ["A", "B"]
        puanlar = [50, 50]

        r = _make_grade_result("ex2", "A", "Ali", ["A", "A"], dogru, puanlar)
        stats = calculate_stats("ex2", [r])
        xlsx_bytes = create_excel("ex2", "Test", [r], stats)

        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["Sonuclar"]

        # Baslik kontrolu
        assert ws.cell(row=1, column=1).value == "Ogrenci Adi"
        assert ws.cell(row=1, column=2).value == "Ogrenci No"
        assert ws.cell(row=1, column=3).value == "Grup"
        assert "S1" in ws.cell(row=1, column=4).value
        assert "S2" in ws.cell(row=1, column=5).value

        # Veri satiri
        assert ws.cell(row=2, column=1).value == "Ali"
        assert ws.cell(row=2, column=3).value == "A"

    def test_stats_sheet_content(self):
        dogru = ["A", "B", "C", "D"]
        puanlar = [25, 25, 25, 25]

        r1 = _make_grade_result("ex3", "A", "O1", ["A", "B", "C", "D"], dogru, puanlar)
        r2 = _make_grade_result("ex3", "B", "O2", ["A", "A", "C", "A"], dogru, puanlar)

        stats = calculate_stats("ex3", [r1, r2])
        xlsx_bytes = create_excel("ex3", "Fizik", [r1, r2], stats)

        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["Istatistikler"]

        # Genel bilgiler
        assert ws.cell(row=1, column=1).value == "Sinav"
        assert ws.cell(row=1, column=2).value == "Fizik"
        assert ws.cell(row=2, column=1).value == "Toplam Ogrenci"
        assert ws.cell(row=2, column=2).value == 2
