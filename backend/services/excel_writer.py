"""Excel (.xlsx) olusturma servisi — 2 sheet: Sonuclar + Istatistikler."""

from __future__ import annotations

import io
import logging

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from models.schemas import ExamStats, GradeResult

logger = logging.getLogger(__name__)

# Hucre arka plan renkleri
_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_BOLD = Font(bold=True)
_CENTER = Alignment(horizontal="center", vertical="center")


def create_excel(
    sinav_id: str,
    sinav_adi: str,
    results: list[GradeResult],
    stats: ExamStats,
) -> bytes:
    """Sonuclar ve istatistik iceren 2 sheet'li Excel dosyasi olustur.

    Args:
        sinav_id: Sinav kimlik bilgisi.
        sinav_adi: Sinav adi (baslik icin).
        results: Puanlanmis ogrenci sonuclari.
        stats: Hesaplanmis istatistikler.

    Returns:
        .xlsx dosya icerigi bytes olarak.
    """
    wb = Workbook()

    _write_results_sheet(wb, sinav_adi, results, stats)
    _write_stats_sheet(wb, sinav_adi, stats)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    logger.info("Excel olusturuldu: sinav=%s, ogrenci=%d", sinav_id, len(results))
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Sheet 1 — Sonuclar
# ---------------------------------------------------------------------------


def _write_results_sheet(
    wb: Workbook,
    sinav_adi: str,
    results: list[GradeResult],
    stats: ExamStats,
) -> None:
    """Ogrenci bazli sonuc tablosu."""
    ws = wb.active
    ws.title = "Sonuclar"

    if not results:
        ws.append(["Sonuc bulunamadi"])
        return

    soru_sayisi = len(results[0].soru_detay)

    # Baslik satiri
    headers = ["Ogrenci Adi", "Ogrenci No", "Grup"]
    for det in results[0].soru_detay:
        puan = det.puan
        headers.append(f"S{det.soru_no} ({puan:.0f}p)")
    headers += ["Toplam", "Maks"]

    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER

    # Gruba gore sirala
    sorted_results = sorted(results, key=lambda r: (r.grup, r.ogrenci_ad or ""))

    for r in sorted_results:
        row = [r.ogrenci_ad or "-", r.ogrenci_no or "-", r.grup]
        for det in r.soru_detay:
            if det.durum == "dogru":
                row.append(det.verilen_cevap or "-")
            elif det.durum == "yanlis":
                row.append(det.verilen_cevap or "-")
            elif det.durum == "gecersiz":
                row.append("X")
            else:  # bos
                row.append("-")
        row += [r.toplam_puan, r.maks_puan]
        ws.append(row)

        # Renklendirme
        row_idx = ws.max_row
        for q_idx, det in enumerate(r.soru_detay):
            col = 4 + q_idx  # 1-indexed: A=1, B=2, C=3, S1=4 ...
            cell = ws.cell(row=row_idx, column=col)
            cell.alignment = _CENTER
            if det.durum == "dogru":
                cell.fill = _GREEN
            elif det.durum == "yanlis":
                cell.fill = _RED
            elif det.durum == "gecersiz":
                cell.fill = _YELLOW

    # Ozet satirlari
    last_row = ws.max_row + 2

    ws.cell(row=last_row, column=1, value="Sinif Ortalamasi").font = _BOLD
    ws.cell(row=last_row, column=2, value=round(stats.sinif_ortalamasi, 2))

    ws.cell(row=last_row + 1, column=1, value="En Yuksek").font = _BOLD
    ws.cell(row=last_row + 1, column=2, value=round(stats.en_yuksek, 2))

    ws.cell(row=last_row + 2, column=1, value="En Dusuk").font = _BOLD
    ws.cell(row=last_row + 2, column=2, value=round(stats.en_dusuk, 2))

    # Sutun genislikleri auto-fit
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 30)


# ---------------------------------------------------------------------------
# Sheet 2 — Istatistikler
# ---------------------------------------------------------------------------


def _write_stats_sheet(
    wb: Workbook,
    sinav_adi: str,
    stats: ExamStats,
) -> None:
    """Istatistik tablosu."""
    ws = wb.create_sheet("Istatistikler")

    # Genel istatistikler
    ws.append(["Sinav", sinav_adi])
    ws.append(["Toplam Ogrenci", stats.toplam_ogrenci])
    ws.append(["Sinif Ortalamasi", stats.sinif_ortalamasi])
    ws.append(["Medyan", stats.medyan])
    ws.append(["Standart Sapma", stats.standart_sapma])
    ws.append(["En Yuksek", stats.en_yuksek])
    ws.append(["En Dusuk", stats.en_dusuk])
    ws.append([])

    for row_idx in range(1, 8):
        ws.cell(row=row_idx, column=1).font = _BOLD

    # Grup bazli ortalamalar
    row = ws.max_row + 1
    ws.cell(row=row, column=1, value="Grup Ortalamalari").font = _BOLD
    row += 1
    ws.cell(row=row, column=1, value="Grup").font = _HEADER_FONT
    ws.cell(row=row, column=1).fill = _HEADER_FILL
    ws.cell(row=row, column=2, value="Ortalama").font = _HEADER_FONT
    ws.cell(row=row, column=2).fill = _HEADER_FILL
    for grup, ort in sorted(stats.grup_ortalamalari.items()):
        row += 1
        ws.cell(row=row, column=1, value=grup).alignment = _CENTER
        ws.cell(row=row, column=2, value=ort)

    # Bos satir
    row += 2

    # Soru bazli dogru oranlari
    ws.cell(row=row, column=1, value="Soru Bazli Dogru Oranlari").font = _BOLD
    row += 1
    ws.cell(row=row, column=1, value="Soru").font = _HEADER_FONT
    ws.cell(row=row, column=1).fill = _HEADER_FILL
    ws.cell(row=row, column=2, value="Dogru Orani (%)").font = _HEADER_FONT
    ws.cell(row=row, column=2).fill = _HEADER_FILL
    for i, oran in enumerate(stats.soru_dogru_oranlari):
        row += 1
        ws.cell(row=row, column=1, value=f"S{i + 1}").alignment = _CENTER
        ws.cell(row=row, column=2, value=round(oran * 100, 1))

    # Bos satir
    row += 2

    # En zor 5 soru
    ws.cell(row=row, column=1, value="En Zor 5 Soru").font = _BOLD
    row += 1
    for i, soru_no in enumerate(stats.en_zor_sorular):
        idx = soru_no - 1
        oran = stats.soru_dogru_oranlari[idx] if idx < len(stats.soru_dogru_oranlari) else 0
        ws.cell(row=row, column=1, value=f"S{soru_no}").alignment = _CENTER
        ws.cell(row=row, column=2, value=f"%{round(oran * 100, 1)} dogru")
        row += 1

    # Sutun genislikleri
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 20
